import frappe
import secrets
import json
from frappe import _


def get_current_profile():
    user_email = frappe.session.user
    profile_name = frappe.db.get_value("Beauty User Profile", {"user": user_email}, "name")
    if not profile_name:
        frappe.throw(_("Beauty profile not found"))
    return profile_name


def authenticate_partner():
    api_key = frappe.get_request_header("X-API-Key")
    api_secret = frappe.get_request_header("X-API-Secret")

    if not api_key or not api_secret:
        frappe.throw(_("Missing API credentials"), frappe.AuthenticationError)

    partner = frappe.db.get_value(
        "Marketplace Partner",
        {"api_key": api_key, "api_secret": api_secret, "is_active": 1, "status": "Active"},
        "name"
    )
    if not partner:
        frappe.throw(_("Invalid or inactive API credentials"), frappe.AuthenticationError)

    return partner


# ---- Partner-facing APIs ----

@frappe.whitelist(allow_guest=True)
def partner_register(data):
    if isinstance(data, str):
        data = frappe.parse_json(data)

    if not data.get("company_name") or not data.get("contact_email"):
        frappe.throw(_("Company name and contact email are required"))

    existing = frappe.db.exists("Marketplace Partner", {"company_name": data["company_name"]})
    if existing:
        frappe.throw(_("A partner with this company name already exists"))

    doc = frappe.get_doc({
        "doctype": "Marketplace Partner",
        "company_name": data["company_name"],
        "contact_email": data.get("contact_email"),
        "contact_person": data.get("contact_person"),
        "webhook_url": data.get("webhook_url"),
        "commission_percent": data.get("commission_percent", 0),
        "integration_type": data.get("integration_type", "Custom API"),
        "country": data.get("country"),
        "status": "Pending",
        "is_active": 0,
    })
    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {
        "message": _("Registration submitted for approval"),
        "partner_id": doc.name,
        "api_key": doc.api_key,
        "api_secret": doc.api_secret,
    }


@frappe.whitelist()
def partner_get_orders():
    partner = authenticate_partner()

    orders = frappe.get_all(
        "Marketplace Order",
        filters={"partner": partner},
        fields=["name", "order_id", "user", "total_price", "order_status", "partner_order_ref", "ordered_date", "delivery_address", "payment_status", "notes"],
        order_by="ordered_date desc"
    )

    for o in orders:
        items = frappe.get_all(
            "Marketplace Order Item",
            filters={"parent": o.name},
            fields=["product", "quantity", "unit_price", "total"]
        )
        for item in items:
            product_info = frappe.db.get_value("Beauty Product", item.product, ["product_name", "brand"], as_dict=True)
            item["product_name"] = product_info.product_name if product_info else None
            item["brand"] = product_info.brand if product_info else None
        o["items"] = items

    return orders


@frappe.whitelist()
def partner_update_order_status(order_id, status, invoice_url=None):
    partner = authenticate_partner()

    valid_statuses = ["Confirmed", "Processing", "Shipped", "Delivered", "Cancelled"]
    if status not in valid_statuses:
        frappe.throw(_("Invalid status. Must be one of: {0}").format(", ".join(valid_statuses)))

    order_name = frappe.db.get_value("Marketplace Order", {"order_id": order_id, "partner": partner}, "name")
    if not order_name:
        frappe.throw(_("Order not found for this partner"))

    doc = frappe.get_doc("Marketplace Order", order_name)
    doc.order_status = status
    if invoice_url:
        doc.invoice_url = invoice_url
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {"message": _("Order status updated to {0}").format(status), "order_id": order_id, "status": status}


@frappe.whitelist()
def partner_get_products():
    partner = authenticate_partner()

    products = frappe.get_all(
        "Beauty Product",
        filters={"partner": partner, "is_active": 1},
        fields=["name", "product_name", "brand", "description", "price", "category", "routine_step", "product_score", "creation"],
        order_by="creation desc"
    )

    return products


@frappe.whitelist()
def partner_create_product(data):
    partner = authenticate_partner()

    if isinstance(data, str):
        data = frappe.parse_json(data)

    if not data.get("product_name") or not data.get("brand"):
        frappe.throw(_("Product name and brand are required"))

    if data.get("price") is not None and float(data.get("price")) < 0:
        frappe.throw(_("Price must be non-negative"))

    existing = frappe.db.exists("Beauty Product", {"product_name": data["product_name"]})
    if existing:
        frappe.throw(_("A product with this name already exists"))

    partner_name = frappe.db.get_value("Marketplace Partner", partner, "company_name")

    doc = frappe.get_doc({
        "doctype": "Beauty Product",
        "product_name": data["product_name"],
        "brand": data.get("brand", partner_name),
        "partner": partner,
        "description": data.get("description", ""),
        "price": data.get("price", 0),
        "category": data.get("category", "Skincare"),
        "routine_step": data.get("routine_step"),
        "product_score": data.get("product_score", 0),
        "is_active": 1,
    })

    if data.get("concerns"):
        for c in data["concerns"]:
            doc.append("concerns", {"concern": c})

    if data.get("ingredients"):
        for i in data["ingredients"]:
            doc.append("ingredients", {"concern": i})

    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {
        "message": _("Product created successfully"),
        "product": doc.as_dict(),
    }


@frappe.whitelist()
def partner_update_product(product_name, data):
    partner = authenticate_partner()

    if isinstance(data, str):
        data = frappe.parse_json(data)

    product_name = frappe.db.get_value(
        "Beauty Product",
        {"product_name": product_name, "partner": partner, "is_active": 1},
        "name"
    )
    if not product_name:
        frappe.throw(_("Product not found or not owned by your company"))

    doc = frappe.get_doc("Beauty Product", product_name)

    updatable_fields = ["brand", "description", "price", "category", "routine_step", "product_score"]
    for field in updatable_fields:
        if field in data:
            setattr(doc, field, data[field])

    if data.get("is_active") is not None:
        doc.is_active = 1 if data["is_active"] else 0

    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "message": _("Product updated successfully"),
        "product": doc.as_dict(),
    }


@frappe.whitelist()
def partner_delete_product(product_name):
    partner = authenticate_partner()

    product_name = frappe.db.get_value(
        "Beauty Product",
        {"product_name": product_name, "partner": partner},
        "name"
    )
    if not product_name:
        frappe.throw(_("Product not found or not owned by your company"))

    doc = frappe.get_doc("Beauty Product", product_name)
    doc.is_active = 0
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {"message": _("Product deactivated successfully")}


@frappe.whitelist()
def partner_bulk_import():
    partner = authenticate_partner()

    from frappe.handler import upload_file
    file_doc = upload_file()

    if not file_doc:
        frappe.throw(_("No file uploaded"))

    file_path = file_doc.get_full_path()
    if not file_path:
        frappe.throw(_("Could not read uploaded file"))

    rows = []
    headers = ["product_name", "brand", "description", "price", "category", "routine_step"]

    if file_path.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            file_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(file_headers, row)))
            wb.close()
        except Exception:
            frappe.throw(_("Invalid Excel file format"))
    else:
        try:
            import csv
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        except Exception:
            frappe.throw(_("Invalid CSV file format"))

    created = 0
    errors = []
    partner_name = frappe.db.get_value("Marketplace Partner", partner, "company_name")

    for i, row in enumerate(rows):
        try:
            name = (row.get("product_name") or "").strip()
            brand = (row.get("brand") or "").strip()

            if not name or not brand:
                errors.append(f"Row {i + 2}: Missing product_name or brand")
                continue

            doc = frappe.get_doc({
                "doctype": "Beauty Product",
                "product_name": name,
                "brand": brand,
                "partner": partner,
                "description": (row.get("description") or "").strip(),
                "price": float(row.get("price") or 0),
                "category": (row.get("category") or "Skincare").strip(),
                "routine_step": (row.get("routine_step") or "").strip(),
                "product_score": float(row.get("product_score") or 0),
                "is_active": 1,
            })
            doc.insert(ignore_permissions=True)
            created += 1
        except Exception as e:
            errors.append(f"Row {i + 2}: {str(e)}")

    frappe.db.commit()

    return {
        "message": _("{0} products created, {1} errors").format(created, len(errors)),
        "created": created,
        "errors": errors,
    }


@frappe.whitelist()
def partner_download_template():
    partner = authenticate_partner()

    headers = ["product_name", "brand", "description", "price", "category", "routine_step"]
    categories = "Skincare,Haircare,Body,Fragrance"
    routine_steps = "Cleanser,Toner,Serum,Moisturizer,SPF,Shampoo,Conditioner,Hair Mask,Hair Serum,Treatment"

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Template"

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="C4A44A", end_color="C4A44A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D1C4A9"),
            right=Side(style="thin", color="D1C4A9"),
            top=Side(style="thin", color="D1C4A9"),
            bottom=Side(style="thin", color="D1C4A9"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 18

        dv_category = DataValidation(type="list", formula1=f'"{categories}"', allow_blank=True)
        dv_category.error = "Please select a valid category"
        dv_category.errorTitle = "Invalid Category"
        ws.add_data_validation(dv_category)
        dv_category.add(f"E2:E1000")

        dv_step = DataValidation(type="list", formula1=f'"{routine_steps}"', allow_blank=True)
        dv_step.error = "Please select a valid routine step"
        dv_step.errorTitle = "Invalid Routine Step"
        ws.add_data_validation(dv_step)
        dv_step.add(f"F2:F1000")

        ws.sheet_properties.tabColor = "C4A44A"

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        frappe.response["filename"] = "aura_product_template.xlsx"
        frappe.response["filecontent"] = output.getvalue()
        frappe.response["type"] = "binary"
        frappe.response["content_type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    except ImportError:
        import csv
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerow(["Example Hydrating Serum", "Your Brand", "A rich hydrating serum with hyaluronic acid", "52.00", "Skincare", "Serum"])
        writer.writerow(["Example Night Cream", "Your Brand", "Retinol night recovery cream", "68.00", "Skincare", "Moisturizer"])

        frappe.response["filename"] = "aura_product_template.csv"
        frappe.response["filecontent"] = output.getvalue()
        frappe.response["type"] = "binary"
        frappe.response["content_type"] = "text/csv"


@frappe.whitelist()
def partner_get_analytics():
    partner = authenticate_partner()

    total_products = frappe.db.count("Beauty Product", {"partner": partner, "is_active": 1})
    total_orders = frappe.db.count("Marketplace Order", {"partner": partner})
    total_revenue = frappe.db.sql("""
        SELECT COALESCE(SUM(total_price), 0) FROM `tabMarketplace Order`
        WHERE partner = %s AND order_status NOT IN ('Cancelled')
    """, partner)[0][0]

    orders_by_status = frappe.db.sql("""
        SELECT order_status, COUNT(*) as count
        FROM `tabMarketplace Order`
        WHERE partner = %s
        GROUP BY order_status
    """, partner, as_dict=True)

    return {
        "total_products": total_products,
        "total_orders": total_orders,
        "total_revenue": total_revenue,
        "orders_by_status": orders_by_status,
    }


@frappe.whitelist()
def partner_get_api_credentials():
    partner = authenticate_partner()

    doc = frappe.get_doc("Marketplace Partner", partner)

    return {
        "api_key": doc.api_key,
        "api_secret": doc.api_secret,
        "webhook_url": doc.webhook_url,
        "company_name": doc.company_name,
        "integration_type": doc.integration_type,
    }


# ---- User-facing APIs ----

@frappe.whitelist()
def place_order(product_id, quantity=1):
    profile = get_current_profile()

    if not frappe.db.exists("Beauty Product", product_id):
        frappe.throw(_("Product not found"))

    product = frappe.get_doc("Beauty Product", product_id)

    partners = frappe.get_all(
        "Marketplace Partner",
        filters={"is_active": 1, "status": "Active"},
        fields=["name"],
        limit=1
    )
    if not partners:
        frappe.throw(_("No active marketplace partner available"))

    doc = frappe.get_doc({
        "doctype": "Marketplace Order",
        "partner": partners[0].name,
        "user": profile,
        "order_status": "Pending",
        "payment_status": "Pending",
        "ordered_date": frappe.utils.now(),
        "delivery_address": frappe.db.get_value("Beauty User Profile", profile, "country") or "",
    })
    doc.append("items", {
        "product": product_id,
        "quantity": int(quantity),
        "unit_price": product.price or 0,
    })
    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {
        "message": _("Order placed successfully"),
        "order_id": doc.order_id,
        "status": doc.order_status,
    }


@frappe.whitelist()
def get_order_history():
    profile = get_current_profile()

    orders = frappe.get_all(
        "Marketplace Order",
        filters={"user": profile},
        fields=["name", "order_id", "partner", "total_price", "order_status", "payment_status", "ordered_date"],
        order_by="ordered_date desc"
    )

    for o in orders:
        partner_name = frappe.db.get_value("Marketplace Partner", o.partner, "company_name")
        o["partner_name"] = partner_name

        items = frappe.get_all(
            "Marketplace Order Item",
            filters={"parent": o.name},
            fields=["product", "quantity", "unit_price", "total"]
        )
        for item in items:
            product_info = frappe.db.get_value("Beauty Product", item.product, ["product_name", "brand"], as_dict=True)
            item["product_name"] = product_info.product_name if product_info else None
            item["brand"] = product_info.brand if product_info else None
        o["items"] = items

    return orders


@frappe.whitelist()
def track_order(order_id):
    profile = get_current_profile()

    order = frappe.db.get_value(
        "Marketplace Order",
        {"order_id": order_id, "user": profile},
        ["name", "order_id", "partner", "total_price", "order_status", "payment_status", "ordered_date", "delivery_address", "invoice_url", "notes"],
        as_dict=True
    )
    if not order:
        frappe.throw(_("Order not found"))

    partner_name = frappe.db.get_value("Marketplace Partner", order.partner, "company_name")
    order["partner_name"] = partner_name

    items = frappe.get_all(
        "Marketplace Order Item",
        filters={"parent": order.name},
        fields=["product", "quantity", "unit_price", "total"]
    )
    for item in items:
        product_info = frappe.db.get_value("Beauty Product", item.product, ["product_name", "brand"], as_dict=True)
        item["product_name"] = product_info.product_name if product_info else None
        item["brand"] = product_info.brand if product_info else None
    order["items"] = items

    return order


# ---- Admin APIs ----

@frappe.whitelist()
def get_marketplace_stats():
    total_partners = frappe.db.count("Marketplace Partner", {"is_active": 1})
    total_orders = frappe.db.count("Marketplace Order")
    pending_orders = frappe.db.count("Marketplace Order", {"order_status": "Pending"})
    completed_orders = frappe.db.count("Marketplace Order", {"order_status": "Delivered"})
    total_revenue = frappe.db.sql("""
        SELECT COALESCE(SUM(total_price), 0) FROM `tabMarketplace Order`
        WHERE order_status NOT IN ('Cancelled')
    """)[0][0]

    orders_by_status = frappe.db.sql("""
        SELECT order_status, COUNT(*) as count
        FROM `tabMarketplace Order`
        GROUP BY order_status
    """, as_dict=True)

    top_partners = frappe.db.sql("""
        SELECT p.company_name, COUNT(o.name) as order_count, COALESCE(SUM(o.total_price), 0) as revenue
        FROM `tabMarketplace Order` o
        JOIN `tabMarketplace Partner` p ON p.name = o.partner
        WHERE o.order_status NOT IN ('Cancelled')
        GROUP BY o.partner
        ORDER BY revenue DESC
        LIMIT 10
    """, as_dict=True)

    return {
        "total_active_partners": total_partners,
        "total_orders": total_orders,
        "pending_orders": pending_orders,
        "completed_orders": completed_orders,
        "total_revenue": total_revenue,
        "orders_by_status": orders_by_status,
        "top_partners": top_partners,
    }


@frappe.whitelist()
def get_partner_performance(partner_id):
    if not frappe.db.exists("Marketplace Partner", partner_id):
        frappe.throw(_("Partner not found"))

    total_orders = frappe.db.count("Marketplace Order", {"partner": partner_id})
    total_revenue = frappe.db.sql("""
        SELECT COALESCE(SUM(total_price), 0) FROM `tabMarketplace Order`
        WHERE partner = %s AND order_status NOT IN ('Cancelled')
    """, partner_id)[0][0]

    orders_by_status = frappe.db.sql("""
        SELECT order_status, COUNT(*) as count
        FROM `tabMarketplace Order`
        WHERE partner = %s
        GROUP BY order_status
    """, partner_id, as_dict=True)

    recent_orders = frappe.get_all(
        "Marketplace Order",
        filters={"partner": partner_id},
        fields=["order_id", "order_status", "total_price", "ordered_date"],
        order_by="ordered_date desc",
        limit=20
    )

    partner = frappe.get_doc("Marketplace Partner", partner_id)

    return {
        "partner": partner.as_dict(),
        "total_orders": total_orders,
        "total_revenue": total_revenue,
        "orders_by_status": orders_by_status,
        "recent_orders": recent_orders,
    }
